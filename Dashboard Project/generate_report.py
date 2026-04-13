#!/usr/bin/env python3
"""
C2 Performance Report Generator
================================
Run each morning after extracting C2-Performance.xlsx from the database.

Output file: C2_Performance_Report_DDMMYYYY.xlsx
Tabs generated:
  {Client}-(DD-MM-YYYY)   -- Raw data per client
  {Client}-Overall        -- DC x service_type performance per client
  C2-Overall              -- All clients combined performance
  C2-PDF                  -- Breach breakdown by DC x service_type (all clients)
  C2-Pivot                -- Attribution % trend for non-AJIO clients
  1+D Eligible Hubs       -- Hub reference list
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os
from datetime import date, datetime
from collections import defaultdict

# ── File paths ─────────────────────────────────────────────────────────────
INPUT_FILE       = "C2-Performance.xlsx"
ELIGIBLE_HUBS_FILE = "1+D Eligible.xlsx"
HISTORY_FILE     = "performance_history.json"   # rolling 8-day trend store

# ── Business constants ─────────────────────────────────────────────────────
SERVICE_TYPES = [
    'Air- Intercity',
    'Intracity NDD',
    'Intracity SDD',
    'Zonal + Air- Intercity',
    'Zonal NDD',
]

ATTRIBUTION_CATEGORIES = [
    'No Breach', 'ODC Connection miss', '1st MR miss', 'DDC Connection miss',
    'JIT/AD miss', 'AH-Intransit', 'Air offload', 'Surface Tagging',
    'Retrieval Delay', 'Hub Capping', 'RTO', 'Pending LM Inscan', '1+ Day Eligible',
]

# C2-PDF: four sections placed side-by-side horizontally
C2_PDF_SECTIONS = [
    {
        'name': 'C2-Air',
        'service_types': ['Air- Intercity'],
        'breach_cols': [
            'No Breach', 'ODC Connection miss', 'DDC Connection miss', 'AH-Intransit',
            'Air offload', 'Surface Tagging', 'MKT Breach', 'Retrieval Delay',
            'Non Eligible', 'RTO', 'Pending LM Inscan', '1+ Day Eligible',
        ],
    },
    {
        'name': 'C2-Zonal',
        'service_types': ['Zonal NDD', 'Zonal + Air- Intercity'],
        'breach_cols': [
            'No Breach', 'ODC Connection miss', 'DDC Connection miss', 'JIT/AD miss',
            'MKT Breach', 'Non Eligible', 'RTO', 'Pending LM Inscan', '1+ Day Eligible',
        ],
    },
    {
        'name': 'Intracity SDD',
        'service_types': ['Intracity SDD'],
        'breach_cols': [
            'No Breach', 'DDC Connection miss', 'Pending LM Inscan', '1+ Day Eligible',
        ],
    },
    {
        'name': 'Intracity NDD',
        'service_types': ['Intracity NDD'],
        'breach_cols': [
            'No Breach', '1st MR miss', 'DDC Connection miss', 'MKT Breach',
            'Post Cutoff', 'RTO', 'Pending LM Inscan',
        ],
    },
]

# Client whose data gets dedicated tabs (others go to C2-Pivot)
PRIMARY_CLIENT_KEYWORD = 'AJIO'

# ── Styling helpers ────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size)

def _border():
    thin = Side(style='thin', color='BFBFBF')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

FILL_DARK_BLUE  = _fill('1F497D')
FILL_MED_BLUE   = _fill('4472C4')
FILL_LIGHT_BLUE = _fill('BDD7EE')
FILL_ORANGE     = _fill('F4B942')
FILL_GREEN      = _fill('70AD47')
FILL_YELLOW     = _fill('FFF2CC')
FILL_GREY       = _fill('D9D9D9')

FONT_WHITE_BOLD = _font(bold=True, color='FFFFFF')
FONT_BOLD       = _font(bold=True)
FONT_NORMAL     = _font()


def style_cell(cell, value=None, fill=None, font=None, align=None, border=True, num_format=None):
    if value is not None:
        cell.value = value
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = _border()
    if num_format:
        cell.number_format = num_format


def auto_col_width(ws, min_width=8, max_width=25):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ── Data loading ───────────────────────────────────────────────────────────

def load_inputs():
    """Load raw shipment data and eligible hubs list."""
    df = pd.read_excel(INPUT_FILE, sheet_name='Sheet1')

    wb_h = openpyxl.load_workbook(ELIGIBLE_HUBS_FILE, read_only=True, data_only=True)
    ws_h = wb_h.active
    eligible_hubs = [
        row[0] for row in ws_h.iter_rows(min_row=2, values_only=True) if row[0]
    ]
    wb_h.close()

    return df, eligible_hubs


def get_report_date(df) -> date:
    dates = df['eligible_attempt_date'].dropna()
    if not dates.empty:
        d = dates.iloc[0]
        return d.date() if hasattr(d, 'date') else d
    return date.today()


def format_client_name(raw_name: str) -> str:
    """AJIO_EXPRESS -> Ajio  |  NYKAA_FASHION -> Nykaa"""
    return raw_name.split('_')[0].capitalize()


# ── Metric computation ─────────────────────────────────────────────────────

def compute_performance(df) -> dict:
    """
    Returns nested dict:
      result[origin_dc][service_type] = {'Vol': int, 'Perf': float}
    """
    result = defaultdict(dict)
    for svc in SERVICE_TYPES:
        svc_df = df[df['service_type'] == svc]
        if svc_df.empty:
            continue
        grp = svc_df.groupby('origin_dc').agg(
            Vol=('awb_number', 'count'),
            NB=('Attribution', lambda x: (x == 'No Breach').sum()),
        )
        for dc, row in grp.iterrows():
            result[dc][svc] = {
                'Vol':  int(row['Vol']),
                'Perf': row['NB'] / row['Vol'] if row['Vol'] else 0.0,
            }
    return dict(result)


def compute_breach_breakdown(df) -> dict:
    """
    Returns nested dict:
      result[section_name][origin_dc] = {'Grand Total': int, 'Perf': float, attr: float, ...}
    """
    result = defaultdict(lambda: defaultdict(dict))
    for section in C2_PDF_SECTIONS:
        sec_name = section['name']
        sec_df = df[df['service_type'].isin(section['service_types'])]
        if sec_df.empty:
            continue
        for dc, grp in sec_df.groupby('origin_dc'):
            total = len(grp)
            perf = (grp['Attribution'] == 'No Breach').sum() / total if total else 0
            result[sec_name][dc] = {'Grand Total': total, 'Perf': perf}
            for attr in section['breach_cols']:
                count = (grp['Attribution'] == attr).sum()
                result[sec_name][dc][attr] = count / total if total else 0
    return {k: dict(v) for k, v in result.items()}


def compute_pivot_data(df) -> dict:
    """
    Returns:
      {service_type: {attribution: pct, ...}, ...}
    for the Attribution breakdown half of the Pivot sheet.
    """
    result = {}
    for svc in SERVICE_TYPES:
        svc_df = df[df['service_type'] == svc]
        if svc_df.empty:
            continue
        total = len(svc_df)
        row = {'Vol': total}
        for attr in ATTRIBUTION_CATEGORIES:
            count = (svc_df['Attribution'] == attr).sum()
            row[attr] = count / total if total else 0
        result[svc] = row
    return result


# ── History management (D-1 to D-8 trend) ─────────────────────────────────

def load_history() -> list:
    """Load rolling performance history. Returns list ordered newest first."""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r') as f:
            return json.load(f)
    return []


def save_history(history: list, report_date: date, pivot_data: dict):
    """Prepend today's performance and keep last 8 days."""
    entry = {
        'date': str(report_date),
        'pivot': pivot_data,
    }
    # Replace same-date entry if re-run, then prepend
    history = [h for h in history if h['date'] != str(report_date)]
    history = [entry] + history
    history = history[:8]   # keep 8 days max
    with open(HISTORY_FILE, 'w') as f:
        json.dump(history, f, indent=2, default=str)
    return history


# ── Sheet writers ──────────────────────────────────────────────────────────

def write_raw_data_tab(wb, df, tab_name):
    """Write raw data with header row styled."""
    ws = wb.create_sheet(tab_name)

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci)
        style_cell(c, value=h, fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD,
                   align=_align('center', 'center'))

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci).value = val

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 20
    # Keep col widths narrow for raw data (many columns)
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    return ws


def write_overall_tab(wb, tab_name, title, perf_data: dict, report_date_str: str):
    """
    Write DC x service_type performance summary.
    Layout matches the Ajio-Overall / C2-Overall example.
    """
    ws = wb.create_sheet(tab_name)

    # ── Row 1: title + service type headers (merged pairs) ────────────────
    style_cell(ws.cell(1, 1), value=f'{title}\n({report_date_str})',
               fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD,
               align=_align('center', 'center', wrap=True))
    ws.row_dimensions[1].height = 28

    col = 2
    svc_col_map = {}
    for svc in SERVICE_TYPES:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        style_cell(ws.cell(1, col), value=svc,
                   fill=FILL_MED_BLUE, font=FONT_WHITE_BOLD,
                   align=_align('center', 'center', wrap=True))
        svc_col_map[svc] = col
        col += 2

    # Grand Total header
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    style_cell(ws.cell(1, col), value='Grand Total',
               fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD,
               align=_align('center', 'center'))
    gt_col = col

    # ── Row 2: Vol / Perfo subheaders ─────────────────────────────────────
    style_cell(ws.cell(2, 1), value='DC',
               fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
    for svc, sc in svc_col_map.items():
        style_cell(ws.cell(2, sc),     value='Vol',   fill=FILL_LIGHT_BLUE, font=FONT_BOLD, align=_align('center'))
        style_cell(ws.cell(2, sc + 1), value='Perfo', fill=FILL_LIGHT_BLUE, font=FONT_BOLD, align=_align('center'))
    style_cell(ws.cell(2, gt_col),     value='Overall Vol',   fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
    style_cell(ws.cell(2, gt_col + 1), value='Overall Perfo', fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))

    # ── Data rows ─────────────────────────────────────────────────────────
    all_dcs = sorted(perf_data.keys())
    for ri, dc in enumerate(all_dcs, start=3):
        style_cell(ws.cell(ri, 1), value=dc, align=_align('left'))
        total_vol = total_nb = 0
        for svc in SERVICE_TYPES:
            if svc in perf_data.get(dc, {}):
                sc = svc_col_map[svc]
                vol  = perf_data[dc][svc]['Vol']
                perf = perf_data[dc][svc]['Perf']
                style_cell(ws.cell(ri, sc),     value=vol,  align=_align('center'))
                style_cell(ws.cell(ri, sc + 1), value=perf, align=_align('center'),
                           num_format='0.00%')
                total_vol += vol
                total_nb  += int(vol * perf)

        style_cell(ws.cell(ri, gt_col),     value=total_vol, align=_align('center'))
        if total_vol:
            style_cell(ws.cell(ri, gt_col + 1),
                       value=total_nb / total_vol, align=_align('center'), num_format='0.00%')

    # ── Grand Total row ───────────────────────────────────────────────────
    gt_row = len(all_dcs) + 3
    style_cell(ws.cell(gt_row, 1), value='Grand Total',
               fill=FILL_GREY, font=FONT_BOLD, align=_align('left'))
    all_vol_total = all_nb_total = 0
    for svc in SERVICE_TYPES:
        sc  = svc_col_map[svc]
        vol = sum(perf_data[dc][svc]['Vol']  for dc in perf_data if svc in perf_data[dc])
        nb  = sum(int(perf_data[dc][svc]['Vol'] * perf_data[dc][svc]['Perf'])
                  for dc in perf_data if svc in perf_data[dc])
        if vol:
            style_cell(ws.cell(gt_row, sc),     value=vol, fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
            style_cell(ws.cell(gt_row, sc + 1), value=nb / vol, fill=FILL_GREY, font=FONT_BOLD,
                       align=_align('center'), num_format='0.00%')
        all_vol_total += vol
        all_nb_total  += nb

    style_cell(ws.cell(gt_row, gt_col),     value=all_vol_total, fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
    if all_vol_total:
        style_cell(ws.cell(gt_row, gt_col + 1),
                   value=all_nb_total / all_vol_total,
                   fill=FILL_GREY, font=FONT_BOLD, align=_align('center'), num_format='0.00%')

    # Column widths
    ws.column_dimensions['A'].width = 22
    for svc, sc in svc_col_map.items():
        ws.column_dimensions[get_column_letter(sc)].width     = 8
        ws.column_dimensions[get_column_letter(sc + 1)].width = 9
    ws.column_dimensions[get_column_letter(gt_col)].width     = 12
    ws.column_dimensions[get_column_letter(gt_col + 1)].width = 14
    ws.freeze_panes = 'B3'


def write_c2_pdf_tab(wb, breakdown: dict, report_date_str: str):
    """
    Write the C2-PDF breach breakdown sheet.
    Four sections placed horizontally (separated by a blank column).
    Each section: rows = DCs, columns = breach categories.
    """
    ws = wb.create_sheet('C2-PDF')

    section_fills = [FILL_DARK_BLUE, FILL_MED_BLUE, FILL_ORANGE, FILL_GREEN]

    start_col = 1
    for sec_idx, section in enumerate(C2_PDF_SECTIONS):
        sec_name    = section['name']
        breach_cols = section['breach_cols']
        sec_fill    = section_fills[sec_idx % len(section_fills)]

        # Get DCs present in this section
        sec_data = breakdown.get(sec_name, {})
        dcs = sorted(sec_data.keys())
        n_cols = 2 + len(breach_cols)   # DC + Grand Total + breach cols

        # Row 1: section header (merged)
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=start_col + n_cols - 1)
        style_cell(ws.cell(1, start_col), value=sec_name,
                   fill=sec_fill, font=FONT_WHITE_BOLD,
                   align=_align('center', 'center'))

        # Row 2: column headers
        style_cell(ws.cell(2, start_col), value='DC',
                   fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
        style_cell(ws.cell(2, start_col + 1), value='Grand Total',
                   fill=FILL_LIGHT_BLUE, font=FONT_BOLD, align=_align('center'))
        for ci, attr in enumerate(breach_cols, start=start_col + 2):
            fill = FILL_GREEN if attr == 'No Breach' else FILL_LIGHT_BLUE
            style_cell(ws.cell(2, ci), value=attr,
                       fill=fill, font=FONT_BOLD,
                       align=_align('center', 'center', wrap=True))

        # Data rows
        for ri, dc in enumerate(dcs, start=3):
            dc_data = sec_data[dc]
            style_cell(ws.cell(ri, start_col), value=dc, align=_align('left'))
            style_cell(ws.cell(ri, start_col + 1),
                       value=dc_data.get('Grand Total', 0), align=_align('center'))
            for ci, attr in enumerate(breach_cols, start=start_col + 2):
                val = dc_data.get(attr, 0)
                num_fmt = '0.00%' if attr != 'Grand Total' else '0'
                style_cell(ws.cell(ri, ci), value=val,
                           align=_align('center'), num_format=num_fmt)

        # Grand Total row
        gt_row = len(dcs) + 3
        style_cell(ws.cell(gt_row, start_col), value='Grand Total',
                   fill=FILL_GREY, font=FONT_BOLD, align=_align('left'))
        gt_total = sum(d.get('Grand Total', 0) for d in sec_data.values())
        style_cell(ws.cell(gt_row, start_col + 1), value=gt_total,
                   fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
        for ci, attr in enumerate(breach_cols, start=start_col + 2):
            total_attr = sum(d.get('Grand Total', 0) * d.get(attr, 0)
                             for d in sec_data.values())
            val = total_attr / gt_total if gt_total else 0
            style_cell(ws.cell(gt_row, ci), value=val,
                       fill=FILL_GREY, font=FONT_BOLD,
                       align=_align('center'), num_format='0.00%')

        # Column widths for this section
        ws.column_dimensions[get_column_letter(start_col)].width = 22
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 12
        for ci in range(start_col + 2, start_col + n_cols):
            ws.column_dimensions[get_column_letter(ci)].width = 14

        start_col += n_cols + 1   # +1 blank separator column

    ws.freeze_panes = 'B3'
    ws.row_dimensions[2].height = 40


def write_pivot_tab(wb, tab_name: str, pivot_data: dict, history: list, report_date_str: str):
    """
    Write the Pivot sheet.
    Left half:  service_type x Attribution % breakdown
    Right half: service_type x D-1..D-8 performance trend
    """
    ws = wb.create_sheet(tab_name)

    # ── Left half: Attribution breakdown ──────────────────────────────────
    row_offset = 3      # start on row 3 (matching example)

    # Title / header
    style_cell(ws.cell(row_offset - 1, 1),
               value=f'Attribution Breakdown ({report_date_str})',
               fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD,
               align=_align('center', 'center'))
    ws.merge_cells(start_row=row_offset - 1, start_column=1,
                   end_row=row_offset - 1, end_column=len(ATTRIBUTION_CATEGORIES) + 1)

    # Column headers
    style_cell(ws.cell(row_offset, 1), value='service_type',
               fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
    for ci, attr in enumerate(ATTRIBUTION_CATEGORIES, 2):
        fill = FILL_GREEN if attr == 'No Breach' else FILL_LIGHT_BLUE
        style_cell(ws.cell(row_offset, ci), value=attr,
                   fill=fill, font=FONT_BOLD,
                   align=_align('center', 'center', wrap=True))

    ws.row_dimensions[row_offset].height = 40

    # Data rows
    all_vol = all_nb = 0
    grand_row = defaultdict(float)
    grand_total_vol = 0

    for ri, svc in enumerate(SERVICE_TYPES, start=row_offset + 1):
        if svc not in pivot_data:
            continue
        row_data = pivot_data[svc]
        style_cell(ws.cell(ri, 1), value=svc, align=_align('left'))
        for ci, attr in enumerate(ATTRIBUTION_CATEGORIES, 2):
            val = row_data.get(attr, 0)
            num_fmt = '0.00%'
            style_cell(ws.cell(ri, ci), value=val,
                       align=_align('center'), num_format=num_fmt)
            grand_row[attr] += val * row_data.get('Vol', 0)
        grand_total_vol += row_data.get('Vol', 0)

    # Grand Total row
    gt_row = row_offset + len(SERVICE_TYPES) + 1
    style_cell(ws.cell(gt_row, 1), value='Grand Total',
               fill=FILL_GREY, font=FONT_BOLD, align=_align('left'))
    for ci, attr in enumerate(ATTRIBUTION_CATEGORIES, 2):
        val = grand_row[attr] / grand_total_vol if grand_total_vol else 0
        style_cell(ws.cell(gt_row, ci), value=val,
                   fill=FILL_GREY, font=FONT_BOLD,
                   align=_align('center'), num_format='0.00%')

    # ── Right half: D-1 to D-8 trend ──────────────────────────────────────
    trend_start_col = len(ATTRIBUTION_CATEGORIES) + 3   # +2 blank separator

    # Header row
    style_cell(ws.cell(row_offset - 1, trend_start_col),
               value='Performance Trend (D-1 = latest)',
               fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD,
               align=_align('center', 'center'))
    ws.merge_cells(start_row=row_offset - 1, start_column=trend_start_col,
                   end_row=row_offset - 1, end_column=trend_start_col + len(history))

    # Column sub-headers: D-1, D-2, ...
    style_cell(ws.cell(row_offset, trend_start_col), value='service_type',
               fill=FILL_GREY, font=FONT_BOLD, align=_align('center'))
    for di, h in enumerate(history, 1):
        d_date = h.get('date', '')
        style_cell(ws.cell(row_offset, trend_start_col + di),
                   value=f'D-{di}\n({d_date})',
                   fill=FILL_LIGHT_BLUE, font=FONT_BOLD,
                   align=_align('center', 'center', wrap=True))

    # Data rows
    gt_trend = defaultdict(list)
    for ri, svc in enumerate(SERVICE_TYPES, start=row_offset + 1):
        style_cell(ws.cell(ri, trend_start_col), value=svc, align=_align('left'))
        for di, h in enumerate(history, 1):
            h_pivot = h.get('pivot', {})
            svc_data = h_pivot.get(svc, {})
            no_breach = svc_data.get('No Breach', 0)
            style_cell(ws.cell(ri, trend_start_col + di),
                       value=no_breach, align=_align('center'), num_format='0.00%')
            gt_trend[di].append((svc_data.get('Vol', 0), no_breach))

    # Grand Total trend row
    gt_row2 = row_offset + len(SERVICE_TYPES) + 1
    style_cell(ws.cell(gt_row2, trend_start_col), value='Grand Total',
               fill=FILL_GREY, font=FONT_BOLD, align=_align('left'))
    for di in range(1, len(history) + 1):
        pairs = gt_trend[di]
        total_vol = sum(p[0] for p in pairs)
        weighted  = sum(p[0] * p[1] for p in pairs)
        val = weighted / total_vol if total_vol else 0
        style_cell(ws.cell(gt_row2, trend_start_col + di),
                   value=val, fill=FILL_GREY, font=FONT_BOLD,
                   align=_align('center'), num_format='0.00%')

    # Column widths
    ws.column_dimensions['A'].width = 25
    for ci in range(2, len(ATTRIBUTION_CATEGORIES) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions[get_column_letter(trend_start_col)].width = 25
    for di in range(1, len(history) + 1):
        ws.column_dimensions[get_column_letter(trend_start_col + di)].width = 14

    ws.freeze_panes = 'B4'


def write_eligible_hubs_tab(wb, eligible_hubs: list):
    ws = wb.create_sheet('1+D Eligible Hubs')
    style_cell(ws.cell(1, 1), value='1+D Eligible',
               fill=FILL_DARK_BLUE, font=FONT_WHITE_BOLD,
               align=_align('center'))
    for ri, hub in enumerate(eligible_hubs, 2):
        ws.cell(ri, 1).value = hub
    ws.column_dimensions['A'].width = 30


# ── Main orchestrator ──────────────────────────────────────────────────────

def generate_report():
    print("=" * 60)
    print("C2 Performance Report Generator")
    print("=" * 60)

    # 1. Load data
    print("\n[1/6] Loading input files...")
    df, eligible_hubs = load_inputs()
    report_date     = get_report_date(df)
    report_date_str = report_date.strftime('%d-%m-%Y')
    clients         = sorted(df['client_name'].dropna().unique())

    print(f"      Date       : {report_date_str}")
    print(f"      Total rows : {len(df):,}")
    print(f"      Clients    : {', '.join(clients)}")
    print(f"      Hubs (1+D) : {len(eligible_hubs):,}")

    # 2. Compute metrics
    print("\n[2/6] Computing metrics...")
    all_perf_data  = compute_performance(df)
    breach_data    = compute_breach_breakdown(df)

    # Per-client perf
    client_perf = {}
    for client in clients:
        client_perf[client] = compute_performance(df[df['client_name'] == client])

    # Determine pivot clients (non-AJIO)
    pivot_clients = [c for c in clients if PRIMARY_CLIENT_KEYWORD not in c.upper()]
    if pivot_clients:
        pivot_df   = df[df['client_name'].isin(pivot_clients)]
    else:
        pivot_df   = df   # fallback: use all data when no non-AJIO clients yet
    pivot_data = compute_pivot_data(pivot_df)

    # 3. History (D-1..D-8 trend)
    print("\n[3/6] Updating performance history...")
    history = load_history()
    history = save_history(history, report_date, pivot_data)
    print(f"      History entries: {len(history)} days")

    # 4. Build workbook
    print("\n[4/6] Building Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # ── Per-client raw data + overview tabs ───────────────────────────────
    for client in clients:
        client_label = format_client_name(client)
        client_df    = df[df['client_name'] == client].copy()

        # Raw data tab
        raw_tab = f"{client_label}-({report_date_str})"[:31]
        print(f"      + Tab: {raw_tab} ({len(client_df):,} rows)")
        write_raw_data_tab(wb, client_df, raw_tab)

        # Overall summary tab
        overall_tab = f"{client_label}-Overall"[:31]
        print(f"      + Tab: {overall_tab}")
        write_overall_tab(wb, overall_tab,
                          f"{client_label}-Performance",
                          client_perf[client],
                          report_date_str)

    # ── C2-Overall (all clients) ───────────────────────────────────────────
    print("      + Tab: C2-Overall")
    write_overall_tab(wb, 'C2-Overall', 'C2-Performance',
                      all_perf_data, report_date_str)

    # ── C2-PDF ────────────────────────────────────────────────────────────
    print("      + Tab: C2-PDF")
    write_c2_pdf_tab(wb, breach_data, report_date_str)

    # ── C2-Pivot (non-AJIO clients, or all if only AJIO exists) ──────────
    pivot_label = 'C2-Pivot'
    if not pivot_clients:
        pivot_label = f"{format_client_name(clients[0])}-Pivot" if clients else 'C2-Pivot'
    print(f"      + Tab: {pivot_label}")
    write_pivot_tab(wb, pivot_label, pivot_data, history, report_date_str)

    # ── 1+D Eligible Hubs ─────────────────────────────────────────────────
    print("      + Tab: 1+D Eligible Hubs")
    write_eligible_hubs_tab(wb, eligible_hubs)

    # 5. Save
    output_file = f"C2_Performance_Report_{report_date.strftime('%d%m%Y')}.xlsx"
    print(f"\n[5/6] Saving report -> {output_file}")
    wb.save(output_file)

    print(f"\n[6/6] Done. Report saved: {output_file}")
    print("=" * 60)
    return output_file


if __name__ == '__main__':
    # Change working directory to the script's location
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    generate_report()
